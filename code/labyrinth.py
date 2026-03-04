from regions import RegionManager, PolygonRegion, CircleRegion, CircularFractionRegion
import numpy as np
import cv2
import os
from logic import EventLogic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class Labyrinth:
	"""
	Clase abstracta para los diferentes laberintos.
	"""
	def __init__(self, video_path: str, treatment: str, subject_id: str, mace_type: str,
				 regions: list, min_detection_area: int, hitbox_size: int,
				 start_time: float, end_time: float | None = None,
				 kernel_size=5, blur_size=0):

		self.video_path       = video_path
		self.treatment        = treatment
		self.subject_id       = subject_id
		self.mace_type        = mace_type
		self.regions          = regions
		self.start_time       = start_time
		self.end_time         = end_time
		self.min_detection_area = min_detection_area
		self.hitbox_size      = hitbox_size
		self.kernel_size      = kernel_size
		self.blur_size        = blur_size
		self.fps              = 0  # se llena al abrir el video en RunExperiment

		self._validate_inputs()

		self.trajectory_x    = []
		self.trajectory_y    = []
		self.trajectory_time = []
		self.results         = {}

	# ----------------------------------------------------------------------
	# Validaciones generales
	# ----------------------------------------------------------------------

	def _validate_inputs(self):
		if not isinstance(self.video_path, str):
			raise ValueError("La dirección del video (video_path) debe ser una cadena de texto.")
		if not isinstance(self.treatment, str):
			raise ValueError("El tratamiento (treatment) debe ser una cadena de texto.")
		if not isinstance(self.subject_id, str):
			raise ValueError("La identificación del sujeto de experimentación (subject_id) debe ser una cadena de texto.")
		if not isinstance(self.mace_type, str):
			raise ValueError("El tipo de laberinto (mace_type) debe ser una cadena de texto.")
		if not isinstance(self.regions, RegionManager):
			raise ValueError("Las regiones de interés (regions) debe ser un RegionManager.")
		if not isinstance(self.start_time, (int, float)):
			raise ValueError("start_time debe ser un número.")
		if self.end_time is None:
			print("Advertencia: end_time no especificado, se procesará todo el video desde start_time hasta el final.")
		else:
			if not isinstance(self.end_time, (int, float)):
				raise ValueError("end_time debe ser un número o None.")
			if self.start_time < 0 or self.end_time < 0:
				raise ValueError("start_time y end_time deben ser números no negativos.")
			if self.start_time >= self.end_time:
				raise ValueError("start_time debe ser menor que end_time.")
		if not isinstance(self.min_detection_area, int) or self.min_detection_area <= 0:
			raise ValueError("min_detection_area debe ser un entero positivo.")
		if not isinstance(self.hitbox_size, int) or self.hitbox_size <= 0:
			raise ValueError("hitbox_size debe ser un entero positivo.")
		if not isinstance(self.fps, int) or self.fps < 0:
			raise ValueError("fps debe ser un entero positivo.")

	# ----------------------------------------------------------------------
	# Métodos abstractos — cada subclase debe implementarlos
	# ----------------------------------------------------------------------

	def process_frame(self, position, time):
		raise NotImplementedError("process_frame debe ser implementada por cada tipo específico de laberinto.")

	def process_video(self, all_results, all_trajectories, all_video_paths, all_first_frames, all_start_times):
		self.write_results(all_results, all_trajectories, all_video_paths, all_first_frames, all_start_times)
		raise NotImplementedError("process_video debe ser implementada por cada tipo específico de laberinto.")

	# ----------------------------------------------------------------------
	# Métodos comunes a todos los laberintos
	# ----------------------------------------------------------------------

	def get_position(self, position, time):
		"""
		Recoge los datos de posición y tiempo obtenidos en cada frame
		y los almacena en listas para su posterior análisis.

		Parámetros
		----------
		position : list
			Lista de coordenadas (x, y) detectadas. Se usa la última posición.
		time : float
			Tiempo actual en segundos.

		Retorna
		-------
		None
		"""
		if len(position) == 0:
			return  # no se detectó posición en este frame
		x, y = position[-1]
		self.trajectory_x.append(x)
		self.trajectory_y.append(y)
		self.trajectory_time.append(time)

	def get_total_distance(self, start_frame=0, end_frame=None):
		"""
		Calcula la distancia total recorrida en píxeles dentro de un rango de frames.

		Los frames de entrada son absolutos del video. Internamente se convierten
		a índices relativos de la trayectoria, que empieza en start_time * fps.

		Parámetros
		----------
		start_frame : int
			Frame absoluto del video desde el cual empezar (default: 0).
		end_frame : int or None
			Frame absoluto del video hasta el cual contar (default: None = hasta el final).

		Retorna
		-------
		float
			Distancia total en píxeles. Retorna 0.0 si no hay suficientes puntos.
		"""
		trajectory = list(zip(self.trajectory_x, self.trajectory_y))

		if len(trajectory) <= 1:
			return 0.0

		# Convertir frames absolutos a índices relativos de la trayectoria
		recording_start = int(self.start_time * self.fps)

		if end_frame is None:
			end_frame = len(trajectory)
		else:
			end_frame = max(0, end_frame - recording_start)

		start_frame = max(0, start_frame - recording_start)
		end_frame   = min(len(trajectory), end_frame)

		trajectory_filtered = trajectory[start_frame:end_frame]

		total_distance = 0.0
		for i in range(1, len(trajectory_filtered)):
			pt1 = trajectory_filtered[i - 1]
			pt2 = trajectory_filtered[i]
			dx  = pt2[0] - pt1[0]
			dy  = pt2[1] - pt1[1]
			total_distance += np.sqrt(dx * dx + dy * dy)

		return total_distance

	def write_results(self, all_results, all_trajectories, all_video_paths, all_first_frames, all_start_times):
		"""
		Esqueleto común para generar Excel y PNGs de trayectoria.
		Delega el resumen específico a _write_summary() de cada subclase.
		"""
		wb = Workbook()
		wb.remove(wb.active)

		for video_name, events_on_each_region in all_results.items():

			traj_x, traj_y    = all_trajectories[video_name]
			self.trajectory_x = traj_x
			self.trajectory_y = traj_y
			# Restaurar start_time del video correspondiente
			self.start_time = all_start_times[video_name]
   
			ws = wb.create_sheet(title=video_name[:31])

			total_distance       = self.get_total_distance()
			total_recording_time = len(self.trajectory_x) / self.fps

			# Metadatos — igual para todos
			self.write_metadata(ws, video_name)

			# Resumen — cada subclase lo implementa diferente
			self.write_summary(ws, events_on_each_region, total_distance, total_recording_time)

			# Formateo final — igual para todos
			self.apply_sheet_format(ws)

			# PNG de trayectoria — igual para todos
			self.save_trajectory_image(
				video_name, traj_x, traj_y, total_distance,
				all_first_frames[video_name],
				os.path.dirname(all_video_paths[video_name])
			)

		# Imágenes combinadas — fuera del loop
		output_dir = os.path.dirname(list(all_video_paths.values())[0])
		first_frame = next((f for f in all_first_frames.values() if f is not None), None)
		self.save_heatmap_image(all_trajectories, first_frame, output_dir)

		# Guardar Excel en la misma carpeta que los videos
		output_dir = os.path.dirname(list(all_video_paths.values())[0])
		filename   = os.path.join(output_dir, f"results_{self.mace_type}_{self.subject_id}_{self.treatment}.xlsx")
		wb.save(filename)
		print(f"Resultados guardados en: {filename}")

	def write_metadata(self, ws, video_name):
		"""Escribe la sección de metadatos del experimento en la hoja."""
		meta = [
			("Sujeto",         self.subject_id),
			("Tratamiento",    self.treatment),
			("Laberinto",      self.mace_type),
			("Video",          video_name),
			("Start time (s)", self.start_time),
			("End time (s)",   self.end_time if self.end_time else "Hasta el final"),
			("FPS",            self.fps),
		]
		for row in meta:
			ws.append(row)
		ws.append([])  # separador

	def write_summary(self, ws, events_on_each_region, total_distance, total_recording_time):
		"""Escribe la sección de resumen. Cada subclase la implementa diferente."""
		raise NotImplementedError

	def compute_region_metrics(self, enter_frames, exit_frames, enter_times, total_distance, total_recording_time):
		"""
		Calcula las métricas de una región a partir de sus frames de entrada/salida.

		Parámetros
		----------
		enter_frames : list
			Frames absolutos de entrada a la región.
		exit_frames : list
			Frames absolutos de salida de la región.
		enter_times : list of tuples
			Lista de (enter_time, exit_time) en segundos, proveniente de state.events.
			Se usa para calcular la latencia con el timestamp real en vez de recalcular
			desde frame_idx, evitando errores por redondeo de fps.
		total_distance : float
			Distancia total recorrida en toda la grabación.
		total_recording_time : float
			Duración total de la grabación en segundos.

		Retorna
		-------
		dict con: event_list, enter_frames, exit_frames, total_time,
				total_dist, latency, pct_time, pct_distance
		"""
		# Corregir si el sujeto quedó dentro de la región al terminar el video
		if len(enter_frames) == len(exit_frames) + 1:
			exit_frames.append(int(self.fps * (self.end_time or total_recording_time + self.start_time)))

		# Duración y distancia por cada evento de entrada/salida
		event_list = []
		for i in range(len(enter_frames)):
			duration = (exit_frames[i] - enter_frames[i]) / self.fps
			distance = self.get_total_distance(start_frame=enter_frames[i], end_frame=exit_frames[i])
			event_list.append((duration, distance))

		total_time_in_region     = sum(e[0] for e in event_list)
		total_distance_in_region = sum(e[1] for e in event_list)

		# Latencia: usar timestamp real del primer ingreso para evitar errores de redondeo de fps
		latency = (enter_frames[0] / self.fps) - self.start_time if enter_frames else None

		# Porcentaje del tiempo de grabación dentro de la región
		pct_time = (total_time_in_region / total_recording_time * 100) if total_recording_time > 0 else 0

		# Porcentaje de la distancia total dentro de la región
		pct_distance = (total_distance_in_region / total_distance * 100) if total_distance > 0 else 0

		return {
			"event_list"  : event_list,
			"enter_frames": enter_frames,
			"exit_frames" : exit_frames,
			"total_time"  : total_time_in_region,
			"total_dist"  : total_distance_in_region,
			"latency"     : latency,
			"pct_time"    : pct_time,
			"pct_distance": pct_distance,
		}

	def write_event_table(self, ws, metrics):
		"""
		Escribe la tabla de detalle por evento con estilos.

		Parámetros
		----------
		ws : Worksheet
			Hoja de Excel donde escribir.
		metrics : dict
			Resultado de _compute_region_metrics().
		"""
		headers = [
			"Evento #", "Frame entrada", "Tiempo entrada (s)",
			"Frame salida", "Tiempo salida (s)",
			"Duración en región (s)", "Distancia en región (px)",
		]
		ws.append(headers)

		# Estilo encabezados: fondo azul, texto blanco, centrado
		header_row = ws.max_row
		for col in range(1, len(headers) + 1):
			cell           = ws.cell(row=header_row, column=col)
			cell.font      = Font(bold=True, color="FFFFFF")
			cell.fill      = PatternFill("solid", start_color="2F5496")
			cell.alignment = Alignment(horizontal="center")

		# Una fila por cada evento de entrada/salida
		for i, event in enumerate(metrics["event_list"]):
			ws.append([
				i + 1,
				metrics["enter_frames"][i],
				round(metrics["enter_frames"][i] / self.fps, 3),
				metrics["exit_frames"][i],
				round(metrics["exit_frames"][i] / self.fps, 3),
				round(event[0], 3),   # duración dentro de la región
				round(event[1], 2),   # distancia dentro de la región
			])

		# Fila de totales — con fórmulas si hay eventos, con ceros si no hay
		total_row_data = ["TOTAL", "", "", "", ""]
		if len(metrics["event_list"]) > 0:
			first_data_row = header_row + 1
			last_data_row  = ws.max_row
			total_row_data += [
				f"=SUM(F{first_data_row}:F{last_data_row})",  # suma duración
				f"=SUM(G{first_data_row}:G{last_data_row})",  # suma distancia
			]
		else:
			total_row_data += [0, 0]  # no hay eventos, no hay suma

		ws.append(total_row_data)
		total_row = ws.max_row
		for col in range(1, len(headers) + 1):
			cell      = ws.cell(row=total_row, column=col)
			cell.font = Font(bold=True)
			cell.fill = PatternFill("solid", start_color="D9E1F2")

	def apply_sheet_format(self, ws):
		"""Aplica formato final a la hoja: centra contenido y ajusta anchos de columna."""
		# Centrar todo el contenido
		for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
			for cell in row:
				cell.alignment = Alignment(horizontal="center")

		# Ancho de columnas
		col_widths = [10, 16, 20, 14, 18, 24, 26]
		for i, width in enumerate(col_widths, 1):
			ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

	def save_trajectory_image(self, video_name, traj_x, traj_y, total_distance, first_frame, output_dir):
		"""
		Guarda imagen PNG de la trayectoria completa sobre el primer frame del video.

		Parámetros
		----------
		video_name : str
			Nombre del video (sin extensión), usado para el nombre del archivo.
		traj_x, traj_y : list
			Coordenadas de la trayectoria.
		total_distance : float
			Distancia total recorrida en píxeles.
		first_frame : np.ndarray or None
			Primer frame del video como fondo de la imagen.
		output_dir : str
			Carpeta donde guardar el PNG.
		"""
		if first_frame is None or len(traj_x) <= 1:
			return

		background = first_frame.copy()
		trajectory = list(zip(traj_x, traj_y))

		# Dibujar regiones
		for region in self.regions.regions:
			region.draw(background, (0, 255, 0), 2)

		# Dibujar trayectoria completa en magenta
		for i in range(1, len(trajectory)):
			cv2.line(background, trajectory[i-1], trajectory[i], (255, 0, 255), 2)

		# Distancia total en la imagen
		cv2.putText(background, f"Distancia: {total_distance:.0f} px",
					(10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)

		img_path = os.path.join(output_dir, f"trajectory_{video_name}.png")
		cv2.imwrite(img_path, background)
		print(f"Imagen guardada en: {img_path}")
  
	def make_black_canvas(self, first_frame):
		"""
		Crea un canvas negro del mismo tamaño que first_frame.
		Retorna (canvas, height, width).
		"""
		if first_frame is not None:
			height, width = first_frame.shape[:2]
		else:
			height, width = 600, 800  # fallback genérico
		canvas = np.zeros((height, width, 3), dtype=np.uint8)
		return canvas, height, width

	def draw_arena_outline(self, canvas):
		"""
		Dibuja el contorno del área experimental sobre el canvas.
		- Morris Pool: círculo completo de la piscina (en gris oscuro como fondo de arena).
		- Cross Maze: no dibuja nada extra (el contorno son las regiones).
		"""
		pass  # la subclase MorrisPool sobreescribe este método

	def compute_density(self, all_trajectories, height, width):
		"""
		Acumula densidad de paso de todas las trayectorias y la normaliza con gamma.
		Gamma < 1 comprime los valores altos, reduciendo la intensidad visual general
		y reservando el rojo/amarillo solo para zonas muy transitadas.
		Retorna np.ndarray uint8 (0-255), o None si no hay datos.
		"""
		density = np.zeros((height, width), dtype=np.float32)
		for traj_x, traj_y in all_trajectories.values():
			if len(traj_x) <= 1:
				continue
			trajectory = list(zip(traj_x, traj_y))
			for j in range(1, len(trajectory)):
				cv2.line(density, trajectory[j-1], trajectory[j], color=1.0, thickness=1)

		if density.max() == 0:
			print("No hay trayectorias para generar heatmap.")
			return None

		# Blur generoso: difumina y reduce picos de intensidad
		density = cv2.GaussianBlur(density, (31, 31), 0)

		# Normalizar linealmente a 0-255
		density_norm = cv2.normalize(density, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
		return density_norm

	def save_heatmap_image(self, all_trajectories, first_frame, output_dir):
		"""
		Genera un mapa de calor combinando todas las trayectorias sobre fondo negro.

		Las intersecciones y zonas más recorridas aparecen en rojo/amarillo (caliente),
		las menos transitadas en azul/verde (frío). Fondo negro = sin actividad.
		"""
		canvas, height, width = self.make_black_canvas(first_frame)

		# Dibujar el área experimental de fondo (solo Morris Pool lo implementa)
		self.draw_arena_outline(canvas)

		density_norm = self.compute_density(all_trajectories, height, width)
		if density_norm is None:
			return

		heatmap_color = cv2.applyColorMap(density_norm, cv2.COLORMAP_JET)

		# Solo pintar píxeles con actividad real (fondo negro donde no hay trayectoria)
		mask = density_norm > 0
		canvas[mask] = heatmap_color[mask]

		# Dibujar regiones en blanco encima del heatmap
		for region in self.regions.regions:
			region.draw(canvas, (255, 255, 255), 2)

		img_path = os.path.join(output_dir, f"heatmap_{self.mace_type}_{self.subject_id}_{self.treatment}.png")
		cv2.imwrite(img_path, canvas)
		print(f"Mapa de calor guardado en: {img_path}")


# ==========================================================================
# Morris Pool
# ==========================================================================

class MorrisPool(Labyrinth):
	"""
	Clase para el experimento de Piscina de Morris.
	- Región de interés: un cuarto de círculo (cuadrante) donde se encuentra la plataforma.
	- Métricas: trayectoria, distancia dentro/fuera de la región, tiempo en el cuadrante
	  por evento individual y acumulado.
	"""

	def __init__(self, video_path: str, treatment: str, subject_id: str, regions: list,
				 min_detection_area: int, hitbox_size: int, start_time: float,
				 end_time: float | None = None, kernel_size=5, blur_size=0):

		super().__init__(video_path, treatment, subject_id, "MorrisPool", regions,
						 min_detection_area, hitbox_size, start_time, end_time,
						 kernel_size, blur_size)

		self.validate_morris_region()
		self.enter_frame = []
		self.exit_frame  = []
		self.event_list  = []

	# ----------------------------------------------------------------------
	# Validaciones específicas para Morris Pool
	# ----------------------------------------------------------------------

	def validate_morris_region(self):
		# 1) Exactamente una región
		if len(self.regions.regions) != 1:
			raise ValueError(
				"Morris Pool requiere exactamente una región de interés "
				"(el cuadrante donde estuvo la plataforma)."
			)
		region = next(iter(self.regions.regions))

		# 2) Tipo correcto
		if not isinstance(region, CircularFractionRegion):
			raise ValueError("La región de Morris Pool debe ser CircularFractionRegion.")

		# 3) Verificar que sea 1/4 de círculo
		angle_span = (region.angle_end - region.angle_start) % 360
		if not np.isclose(angle_span, 90.0, atol=1e-6):
			raise ValueError("La región de Morris Pool debe ser un cuarto de círculo (90°).")

	# ----------------------------------------------------------------------
	# Visualización específica: círculo de la piscina como fondo
	# ----------------------------------------------------------------------

	def draw_arena_outline(self, canvas):
		"""
		Dibuja el círculo completo de la piscina en gris oscuro sobre el canvas.
		Usa el centro y radio de la única región CircularFractionRegion definida.
		"""
		region = next(iter(self.regions.regions))
		center = tuple(map(int, region.center))
		radius = int(region.radius)
		cv2.circle(canvas, center, radius, (40, 40, 40), -1)
		cv2.circle(canvas, center, radius, (180, 180, 180), 2)

	def save_heatmap_image(self, all_trajectories, first_frame, output_dir):
		"""
		Heatmap específico para Morris Pool:
		- El calor queda recortado dentro del círculo de la piscina.
		- Se dibuja una barra de escala de temperatura (frío→caliente).
		- Fondo negro fuera del círculo.
		"""
		canvas, height, width = self.make_black_canvas(first_frame)

		region = next(iter(self.regions.regions))
		center = tuple(map(int, region.center))
		radius = int(region.radius)

		density_norm = self.compute_density(all_trajectories, height, width)
		if density_norm is None:
			return

		heatmap_color = cv2.applyColorMap(density_norm, cv2.COLORMAP_JET)

		# Máscara circular — solo pintar dentro de la piscina
		circle_mask = np.zeros((height, width), dtype=np.uint8)
		cv2.circle(circle_mask, center, radius, 255, -1)

		# Fondo gris oscuro dentro del círculo, negro fuera
		cv2.circle(canvas, center, radius, (40, 40, 40), -1)

		# Aplicar heatmap solo donde hay densidad Y dentro del círculo
		heat_mask = (density_norm > 0) & (circle_mask == 255)
		canvas[heat_mask] = heatmap_color[heat_mask]

		# Contorno de la piscina y cuadrante en blanco
		cv2.circle(canvas, center, radius, (255, 255, 255), 2)
		for reg in self.regions.regions:
			reg.draw(canvas, (255, 255, 255), 2)

		# --- Barra de escala de temperatura ---
		bar_x      = width - 30    # posición horizontal de la barra
		bar_y_top  = 20            # margen superior
		bar_height = height - 40   # altura de la barra
		bar_width  = 18

		# Gradiente vertical de 0 (frío, abajo) a 255 (caliente, arriba)
		for i in range(bar_height):
			value     = int(255 * (1.0 - i / bar_height))  # arriba=caliente
			color_bar = cv2.applyColorMap(np.array([[value]], dtype=np.uint8), cv2.COLORMAP_JET)[0][0]
			cv2.rectangle(
				canvas,
				(bar_x, bar_y_top + i),
				(bar_x + bar_width, bar_y_top + i + 1),
				color_bar.tolist(), -1
			)

		# Borde de la barra
		cv2.rectangle(canvas, (bar_x, bar_y_top), (bar_x + bar_width, bar_y_top + bar_height), (255, 255, 255), 1)

		# Etiquetas "Alto" y "Bajo"
		font       = cv2.FONT_HERSHEY_SIMPLEX
		font_scale = 0.45
		thickness  = 1
		cv2.putText(canvas, "Alto", (bar_x - 2, bar_y_top - 5),       font, font_scale, (255, 255, 255), thickness)
		cv2.putText(canvas, "Bajo", (bar_x - 2, bar_y_top + bar_height + 14), font, font_scale, (255, 255, 255), thickness)

		img_path = os.path.join(output_dir, f"heatmap_{self.mace_type}_{self.subject_id}_{self.treatment}.png")
		cv2.imwrite(img_path, canvas)
		print(f"Mapa de calor guardado en: {img_path}")

	# ----------------------------------------------------------------------
	# Procesamiento
	# ----------------------------------------------------------------------

	def process_frame(self, position, time):
		"""
		Procesa un frame a la vez para extraer la trayectoria del sujeto.

		Parámetros
		----------
		position : list
			Lista de coordenadas (x, y) detectadas en el frame actual.
		time : float
			Tiempo actual en segundos.

		Retorna
		-------
		None
		"""
		self.get_position(position, time)

	def process_video(self, all_results, all_trajectories, all_video_paths, all_first_frames, all_start_times):
		"""
		Procesa los resultados de todos los videos y genera los outputs finales.

		Parámetros
		----------
		all_results : dict
			{nombre_video: events_on_each_region}
		all_trajectories : dict
			{nombre_video: (trajectory_x, trajectory_y)}
		all_video_paths : dict
			{nombre_video: ruta_absoluta}
		all_first_frames : dict
			{nombre_video: primer_frame}
		all_start_times : dict
			{nombre_video: start_time}
		"""
		self.write_results(all_results, all_trajectories, all_video_paths, all_first_frames, all_start_times)

	# ----------------------------------------------------------------------
	# Métodos auxiliares (se mantienen por documentación histórica)
	# ----------------------------------------------------------------------

	def get_time_index_in_out_of_region(self, events_on_each_region):
		"""
		Obtiene los índices de entrada y salida de la región de interés.
		"""
		region_id        = next(iter(self.regions.regions)).region_id
		self.enter_frame = events_on_each_region[region_id].enter_frame
		self.exit_frame  = events_on_each_region[region_id].exit_frame
		self.check_enter_exit_frame_lists()

	def check_enter_exit_frame_lists(self):
		"""
		Verifica que las listas de entrada y salida tengan la misma longitud.
		Si hay una entrada sin salida, agrega una salida al final del video.
		"""
		if len(self.enter_frame) != len(self.exit_frame):
			print(f"Advertencia: entradas {len(self.enter_frame)} y salidas {len(self.exit_frame)} no coinciden.")
			if len(self.enter_frame) == len(self.exit_frame) + 1:
				print("Hay una entrada más que salidas — se agrega salida al final del video.")
				self.exit_frame.append(int(self.fps * self.end_time))
			elif len(self.enter_frame) > len(self.exit_frame):
				print("Hay más entradas que salidas. REVISAR LÓGICAS.")
			else:
				print("Hay más salidas que entradas. REVISAR LÓGICAS.")

	def get_distance_and_time_inside_region(self):
		"""
		Calcula distancia y tiempo dentro de la región para cada evento.
		"""
		for i in range(len(self.enter_frame)):
			self.event_list.append([])
			self.event_list[i].append((self.exit_frame[i] - self.enter_frame[i]) / self.fps)
			self.event_list[i].append(self.get_total_distance(
				start_frame=self.enter_frame[i], end_frame=self.exit_frame[i]))

	# ----------------------------------------------------------------------
	# Resultados
	# ----------------------------------------------------------------------

	def write_summary(self, ws, events_on_each_region, total_distance, total_recording_time):
		"""Resumen de Morris Pool: una sola región."""
		region_id    = next(iter(self.regions.regions)).region_id
		state        = events_on_each_region[region_id]
		enter_frames = list(state.enter_frame)
		exit_frames  = list(state.exit_frame)
		enter_times  = list(state.events)  # lista de (enter_time, exit_time
		m = self.compute_region_metrics(
			enter_frames, exit_frames, enter_times,
			total_distance, total_recording_time
		)
		ws.append(["RESUMEN", ""])
		ws.append(["Nº de entradas a la región",    len(m["enter_frames"])])
		ws.append(["Tiempo total en región (s)",     round(m["total_time"], 3)])
		ws.append(["Distancia total recorrida (px)", round(total_distance, 2)])
		ws.append(["Latencia al primer ingreso (s)", round(m["latency"], 3) if m["latency"] is not None else "No entró"])
		ws.append(["% tiempo en región",             round(m["pct_time"], 2)])
		ws.append(["% distancia en región",          round(m["pct_distance"], 2)])
		ws.append([])
		self.write_event_table(ws, m)


# ==========================================================================
# Cross Maze
# ==========================================================================

class CrossMaze(Labyrinth):
	"""
	Clase para laberintos en cruz.
	- Soporta cualquier número de regiones poligonales (brazos).
	- Métricas por región: latencia, entradas, tiempo, distancia, porcentajes.
	"""

	def __init__(self, video_path, treatment, subject_id, regions,
				 min_detection_area, hitbox_size, start_time, end_time=None,
				 kernel_size=5, blur_size=0):

		super().__init__(video_path, treatment, subject_id, "CrossMaze", regions,
						 min_detection_area, hitbox_size, start_time, end_time,
						 kernel_size, blur_size)

		self.validate_cross_maze_regions()

	# ----------------------------------------------------------------------
	# Validaciones específicas para Cross Maze
	# ----------------------------------------------------------------------

	def validate_cross_maze_regions(self):
		if len(self.regions.regions) < 2:
			raise ValueError("CrossMaze requiere al menos 2 regiones de interés.")
		for region in self.regions.regions:
			if not isinstance(region, PolygonRegion):
				raise ValueError("Todas las regiones de CrossMaze deben ser PolygonRegion.")

	# ----------------------------------------------------------------------
	# Procesamiento
	# ----------------------------------------------------------------------

	def process_frame(self, position, time):
		self.get_position(position, time)

	def process_video(self, all_results, all_trajectories, all_video_paths, all_first_frames, all_start_times):
		"""
		Procesa los resultados de todos los videos y genera los outputs finales.

		Parámetros
		----------
		all_results : dict
			{nombre_video: events_on_each_region}
		all_trajectories : dict
			{nombre_video: (trajectory_x, trajectory_y)}
		all_video_paths : dict
			{nombre_video: ruta_absoluta}
		all_first_frames : dict
			{nombre_video: primer_frame}
		"""
		self.write_results(all_results, all_trajectories, all_video_paths, all_first_frames, all_start_times)

	# ----------------------------------------------------------------------
	# Resultados
	# ----------------------------------------------------------------------

	def write_summary(self, ws, events_on_each_region, total_distance, total_recording_time):
		"""Resumen de CrossMaze: una sección por región."""
		ws.append(["RESUMEN GLOBAL", ""])
		ws.append(["Distancia total recorrida (px)", round(total_distance, 2)])
		ws.append(["Duración total grabación (s)",   round(len(self.trajectory_x) / self.fps, 3)])
		ws.append([])

		for region in self.regions.regions:
			state = events_on_each_region[region.region_id]
			enter_frames = list(state.enter_frame)
			exit_frames  = list(state.exit_frame)
			enter_times  = list(state.events)  # lista de (enter_time, exit_time)
			m = self.compute_region_metrics(
						enter_frames, exit_frames, enter_times,
						total_distance, len(self.trajectory_x) / self.fps
					)
			ws.append([f"REGIÓN: {region.region_id}", ""])
			ws.append(["Nº de entradas",                len(m["enter_frames"])])
			ws.append(["Tiempo total en región (s)",     round(m["total_time"], 3)])
			ws.append(["Distancia en región (px)",       round(m["total_dist"], 2)])
			ws.append(["Latencia al primer ingreso (s)", round(m["latency"], 3) if m["latency"] is not None else "No entró"])
			ws.append(["% tiempo en región",             round(m["pct_time"], 2)])
			ws.append(["% distancia en región",          round(m["pct_distance"], 2)])
			ws.append([])
			self.write_event_table(ws, m)
			ws.append([])  # separador entre regiones