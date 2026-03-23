"""Clase base para los diferentes tipos de laberinto.

Define la interfaz común y los métodos compartidos entre todas las
implementaciones concretas (CrossMaze, MorrisPool, etc.).
"""

from regions import RegionManager, PolygonRegion, CircleRegion, CircularFractionRegion
import numpy as np
import cv2
import os
from logic import EventLogic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class Labyrinth:
	"""Clase base abstracta para los diferentes laberintos.

	Define la interfaz y los métodos comunes a todos los experimentos:
	lectura de posiciones, cálculo de distancia, generación de Excel e
	imágenes de trayectoria y mapa de calor.

	Las subclases deben implementar ``process_frame``, ``process_video``
	y ``write_summary``.

	Attributes:
		video_path: Ruta al archivo de video o carpeta con videos.
		treatment: Nombre del tratamiento aplicado al sujeto.
		subject_id: Identificador del sujeto de experimentación.
		mace_type: Tipo de laberinto (e.g. ``"CrossMaze"``, ``"MorrisPool"``).
		regions: ``RegionManager`` con las regiones de interés.
		start_time: Tiempo en segundos desde el que se empieza a registrar.
		end_time: Tiempo en segundos en el que termina el registro, o None
			para procesar hasta el final del video.
		min_detection_area: Área mínima en píxeles para considerar un contorno
			como el ratón.
		hitbox_size: Semilado de la hitbox cuadrada en píxeles.
		kernel_size: Tamaño del kernel morfológico del tracker.
		blur_size: Tamaño del kernel de GaussianBlur del tracker. 0 desactiva
			el blur.
		fps: FPS del video. Se asigna al abrir el video en ``RunExperiment``.
		trajectory_x: Lista de coordenadas x de la trayectoria grabada.
		trajectory_y: Lista de coordenadas y de la trayectoria grabada.
		trajectory_time: Lista de timestamps en segundos de cada posición.
		results: Diccionario de resultados (uso interno de subclases).
	"""

	def __init__(
		self,
		video_path: str,
		treatment: str,
		subject_id: str,
		mace_type: str,
		regions: RegionManager,
		min_detection_area: int,
		hitbox_size: int,
		start_time: float,
		end_time: float | None = None,
		kernel_size: int = 5,
		blur_size: int = 0,
	):
		"""Inicializa los parámetros comunes del experimento.

		Args:
			video_path: Ruta al archivo de video o carpeta con videos.
			treatment: Nombre del tratamiento aplicado al sujeto.
			subject_id: Identificador del sujeto de experimentación.
			mace_type: Tipo de laberinto (e.g. ``"CrossMaze"``).
			regions: ``RegionManager`` con las regiones de interés.
			min_detection_area: Área mínima en píxeles para detectar al ratón.
			hitbox_size: Semilado de la hitbox cuadrada en píxeles.
			start_time: Tiempo en segundos desde el que se empieza a registrar.
			end_time: Tiempo en segundos en que termina el registro, o None
				para procesar hasta el final del video.
			kernel_size: Tamaño del kernel morfológico del tracker.
			blur_size: Tamaño del kernel GaussianBlur. 0 desactiva el blur.

		Raises:
			ValueError: Si alguno de los parámetros no cumple las restricciones
				de tipo o valor definidas en ``_validate_inputs``.
		"""
		self.video_path         = video_path
		self.treatment          = treatment
		self.subject_id         = subject_id
		self.mace_type          = mace_type
		self.regions            = regions
		self.start_time         = start_time
		self.end_time           = end_time
		self.min_detection_area = min_detection_area
		self.hitbox_size        = hitbox_size
		self.kernel_size        = kernel_size
		self.blur_size          = blur_size
		self.fps                = 0  # se sobreescribe al abrir el video en RunExperiment

		self._validate_inputs()

		self.trajectory_x    = []
		self.trajectory_y    = []
		self.trajectory_time = []
		self.results         = {}

	# ------------------------------------------------------------------
	# Validaciones generales
	# ------------------------------------------------------------------

	def _validate_inputs(self) -> None:
		"""Valida los parametros del experimento.

		Raises:
			ValueError: Si algun parametro no cumple las restricciones de
				tipo o valor esperadas.
		"""
		if not isinstance(self.video_path, str):
			raise ValueError("video_path debe ser una cadena de texto.")
		if not isinstance(self.treatment, str):
			raise ValueError("treatment debe ser una cadena de texto.")
		if not isinstance(self.subject_id, str):
			raise ValueError("subject_id debe ser una cadena de texto.")
		if not isinstance(self.mace_type, str):
			raise ValueError("mace_type debe ser una cadena de texto.")
		if not isinstance(self.regions, RegionManager):
			raise ValueError("regions debe ser un RegionManager.")
		if not isinstance(self.start_time, (int, float)):
			raise ValueError("start_time debe ser un numero.")
		if self.end_time is None:
			print("Advertencia: end_time no especificado, se procesara el video hasta el final.")
		else:
			if not isinstance(self.end_time, (int, float)):
				raise ValueError("end_time debe ser un numero o None.")
			if self.start_time < 0 or self.end_time < 0:
				raise ValueError("start_time y end_time deben ser numeros no negativos.")
			if self.start_time >= self.end_time:
				raise ValueError("start_time debe ser menor o igual que end_time.")
		if not isinstance(self.min_detection_area, int) or self.min_detection_area <= 0:
			raise ValueError("min_detection_area debe ser un entero positivo.")
		if not isinstance(self.hitbox_size, int) or self.hitbox_size <= 0:
			raise ValueError("hitbox_size debe ser un entero positivo.")
		if not isinstance(self.fps, int) or self.fps < 0:
			raise ValueError("fps debe ser un entero no negativo.")

	# ------------------------------------------------------------------
	# Métodos abstractos — cada subclase debe implementarlos
	# ------------------------------------------------------------------

	def process_frame(self, position, time: float) -> None:
		"""Procesa un frame individual. Debe ser implementado por cada subclase.

		Args:
			position: Lista de coordenadas ``(x, y)`` detectadas en el frame.
			time: Tiempo actual en segundos.

		Raises:
			NotImplementedError: Siempre. La subclase debe sobreescribir este metodo.
		"""
		raise NotImplementedError("process_frame debe ser implementada por cada tipo de laberinto.")

	def process_video(
		self,
		all_results: dict,
		all_trajectories: dict,
		all_video_paths: dict,
		all_first_frames: dict,
		all_start_times: dict,
	) -> None:
		"""Procesa los resultados de todos los videos. Debe ser implementado por cada subclase.

		No llama a ``write_results`` — esa responsabilidad recae completamente
		en la subclase para evitar efectos secundarios si se llama a ``super()``.

		Args:
			all_results: ``{nombre_video: events_on_each_region}``.
			all_trajectories: ``{nombre_video: (trajectory_x, trajectory_y)}``.
			all_video_paths: ``{nombre_video: ruta_absoluta}``.
			all_first_frames: ``{nombre_video: primer_frame}``.
			all_start_times: ``{nombre_video: start_time}``.

		Raises:
			NotImplementedError: Siempre. La subclase debe sobreescribir este metodo.
		"""
		raise NotImplementedError("process_video debe ser implementada por cada tipo de laberinto.")

	# ------------------------------------------------------------------
	# Métodos comunes a todos los laberintos
	# ------------------------------------------------------------------

	def get_position(self, position: tuple | None, time: float) -> None:
		"""Almacena la posición y el tiempo del frame actual en la trayectoria.

		Si la posición es None (no se detectó ratón en el frame), no se guarda nada.

		Args:
			position: Coordenadas ``(x, y)`` del ratón detectado, o None si no
				se detectó.
			time: Tiempo actual en segundos.
		"""
		if position is None:
			return
		x, y = position
		self.trajectory_x.append(x)
		self.trajectory_y.append(y)
		self.trajectory_time.append(time)

	def get_total_distance(self, start_frame: int = 0, end_frame: int = None) -> float:
		"""Calcula la distancia total recorrida en píxeles dentro de un rango de frames.

		Los frames de entrada son absolutos del video. Internamente se convierten
		a índices relativos de la trayectoria, que empieza en ``start_time * fps``.

		Args:
			start_frame: Frame absoluto del video desde el cual empezar.
			end_frame: Frame absoluto del video hasta el cual contar. None
				procesa hasta el último punto de la trayectoria.

		Returns:
			Distancia total en píxeles. Retorna 0.0 si no hay suficientes puntos.
		"""
		trajectory = list(zip(self.trajectory_x, self.trajectory_y))

		if len(trajectory) <= 1:
			return 0.0

		# Convertir frames absolutos a índices relativos de la trayectoria
		recording_start = int(self.start_time * self.fps)

		if end_frame is None:
			end_frame = len(trajectory)
		else:
			end_frame = max(0, end_frame - recording_start)

		start_frame = max(0, start_frame - recording_start)
		end_frame   = min(len(trajectory), end_frame)

		trajectory_filtered = trajectory[start_frame:end_frame]

		total_distance = 0.0
		for i in range(1, len(trajectory_filtered)):
			pt1 = trajectory_filtered[i - 1]
			pt2 = trajectory_filtered[i]
			dx  = pt2[0] - pt1[0]
			dy  = pt2[1] - pt1[1]
			total_distance += np.sqrt(dx * dx + dy * dy)

		return total_distance

	def write_results(
		self,
		all_results: dict,
		all_trajectories: dict,
		all_video_paths: dict,
		all_first_frames: dict,
		all_start_times: dict,
	) -> None:
		"""Genera el archivo Excel y las imágenes PNG de trayectoria y heatmap.

		Itera sobre todos los videos procesados, crea una hoja por video en el
		Excel y delega el contenido específico del resumen a ``write_summary``
		de cada subclase.

		Args:
			all_results: ``{nombre_video: events_on_each_region}``.
			all_trajectories: ``{nombre_video: (trajectory_x, trajectory_y)}``.
			all_video_paths: ``{nombre_video: ruta_absoluta}``.
			all_first_frames: ``{nombre_video: primer_frame}``.
			all_start_times: ``{nombre_video: start_time}``.
		"""
		wb = Workbook()
		wb.remove(wb.active)

		for video_name, events_on_each_region in all_results.items():

			traj_x, traj_y    = all_trajectories[video_name]
			self.trajectory_x = traj_x
			self.trajectory_y = traj_y
			self.start_time   = all_start_times[video_name]

			ws = wb.create_sheet(title=video_name[:31])

			total_distance       = self.get_total_distance()
			total_recording_time = len(self.trajectory_x) / self.fps

			self.write_metadata(ws, video_name)
			self.write_summary(ws, events_on_each_region, total_distance, total_recording_time)
			self.apply_sheet_format(ws)

			self.save_trajectory_image(
				video_name, traj_x, traj_y, total_distance,
				all_first_frames[video_name],
				os.path.dirname(all_video_paths[video_name]),
			)

		# Heatmap combinado con todas las trayectorias
		output_dir  = os.path.dirname(list(all_video_paths.values())[0])
		first_frame = next((f for f in all_first_frames.values() if f is not None), None)
		self.save_heatmap_image(all_trajectories, first_frame, output_dir)

		# Guardar Excel en la misma carpeta que los videos
		filename = os.path.join(
			output_dir,
			f"results_{self.mace_type}_{self.subject_id}_{self.treatment}.xlsx",
		)
		wb.save(filename)
		print(f"Resultados guardados en: {filename}")

	def write_metadata(self, ws, video_name: str) -> None:
		"""Escribe la sección de metadatos del experimento en la hoja de Excel.

		Args:
			ws: Hoja de Excel (``Worksheet``) donde escribir.
			video_name: Nombre del video procesado.
		"""
		meta = [
			("Sujeto",         self.subject_id),
			("Tratamiento",    self.treatment),
			("Laberinto",      self.mace_type),
			("Video",          video_name),
			("Start time (s)", self.start_time),
			("End time (s)",   self.end_time if self.end_time else "Hasta el final"),
			("FPS",            self.fps),
		]
		for row in meta:
			ws.append(row)
		ws.append([])  # separador

	def write_summary(
		self,
		ws,
		events_on_each_region: dict,
		total_distance: float,
		total_recording_time: float,
	) -> None:
		"""Escribe la seccion de resumen en la hoja. Debe ser implementado por cada subclase.

		Args:
			ws: Hoja de Excel (``Worksheet``) donde escribir.
			events_on_each_region: ``{region_id: ZoneState}`` con los eventos
				de cada region.
			total_distance: Distancia total recorrida en pixeles.
			total_recording_time: Duracion total de la grabacion en segundos.

		Raises:
			NotImplementedError: Si la subclase no implementa este metodo.
		"""
		raise NotImplementedError

	def compute_region_metrics(
		self,
		enter_frames: list,
		exit_frames: list,
		enter_times: list,
		total_distance: float,
		total_recording_time: float,
	) -> dict:
		"""Calcula las métricas de una región a partir de sus frames de entrada/salida.

		Filtra eventos con distancia menor a 1 píxel y duración menor a 1 segundo
		antes de calcular totales. Si el sujeto quedó dentro al terminar el video
		se añade una salida artificial al final de la grabación.

		Args:
			enter_frames: Frames absolutos de entrada a la región.
			exit_frames: Frames absolutos de salida de la región.
			enter_times: Lista de tuplas (enter_time, exit_time) en segundos,
				proveniente de ZoneState.events.
			total_distance: Distancia total recorrida en toda la grabación.
			total_recording_time: Duración total de la grabación en segundos.

		Returns:
			Diccionario con las siguientes claves:

			- event_list: lista de tuplas (duración, distancia) por evento.
			- enter_frames: lista de frames de entrada filtrada.
			- exit_frames: lista de frames de salida filtrada.
			- total_time: tiempo total en la región en segundos.
			- total_dist: distancia total dentro de la región en píxeles.
			- latency: latencia al primer ingreso en segundos, o None si no hubo
			ingresos.
			- pct_time: porcentaje del tiempo de grabación dentro de la región.
			- pct_distance: porcentaje de la distancia total dentro de la región.
		"""
		# Corregir si el sujeto quedó dentro de la región al terminar el video
		if len(enter_frames) == len(exit_frames) + 1:
			exit_frames.append(int(self.fps * (self.end_time or total_recording_time + self.start_time)))

		# Construir event_list filtrando entradas falsas
		event_list           = []
		enter_frames_filtered = []
		exit_frames_filtered  = []

		for i in range(len(enter_frames)):
			duration = (exit_frames[i] - enter_frames[i]) / self.fps
			distance = self.get_total_distance(start_frame=enter_frames[i], end_frame=exit_frames[i])
			if distance < 10.0: # Agregar si hace falta  and duration < 1.0
				continue
			event_list.append((duration, distance))
			enter_frames_filtered.append(enter_frames[i])
			exit_frames_filtered.append(exit_frames[i])

		enter_frames = enter_frames_filtered
		exit_frames  = exit_frames_filtered

		total_time_in_region     = sum(e[0] for e in event_list)
		total_distance_in_region = sum(e[1] for e in event_list)

		latency = (enter_frames[0] / self.fps) - self.start_time if enter_frames else None

		pct_time     = (total_time_in_region     / total_recording_time * 100) if total_recording_time > 0 else 0
		pct_distance = (total_distance_in_region / total_distance       * 100) if total_distance       > 0 else 0

		return {
			"event_list"  : event_list,
			"enter_frames": enter_frames,
			"exit_frames" : exit_frames,
			"total_time"  : total_time_in_region,
			"total_dist"  : total_distance_in_region,
			"latency"     : latency,
			"pct_time"    : pct_time,
			"pct_distance": pct_distance,
		}

	def write_event_table(self, ws, metrics: dict) -> None:
		"""Escribe la tabla de detalle por evento con estilos en la hoja de Excel.

		Cada fila corresponde a un evento de entrada/salida. La última fila
		contiene los totales de duración y distancia mediante fórmulas de suma.

		Args:
			ws: Hoja de Excel (``Worksheet``) donde escribir.
			metrics: Diccionario resultado de ``compute_region_metrics``.
		"""
		headers = [
			"Evento #", "Frame entrada", "Tiempo entrada (s)",
			"Frame salida", "Tiempo salida (s)",
			"Duración en región (s)", "Distancia en región (px)",
		]
		ws.append(headers)

		# Estilo encabezados: fondo azul, texto blanco, centrado
		header_row = ws.max_row
		for col in range(1, len(headers) + 1):
			cell           = ws.cell(row=header_row, column=col)
			cell.font      = Font(bold=True, color="FFFFFF")
			cell.fill      = PatternFill("solid", start_color="2F5496")
			cell.alignment = Alignment(horizontal="center")

		# Una fila por evento
		for i, event in enumerate(metrics["event_list"]):
			ws.append([
				i + 1,
				metrics["enter_frames"][i],
				round(metrics["enter_frames"][i] / self.fps, 3),
				metrics["exit_frames"][i],
				round(metrics["exit_frames"][i] / self.fps, 3),
				round(event[0], 3),
				round(event[1], 2),
			])

		# Fila de totales
		total_row_data = ["TOTAL", "", "", "", ""]
		if len(metrics["event_list"]) > 0:
			first_data_row = header_row + 1
			last_data_row  = ws.max_row
			total_row_data += [
				f"=SUM(F{first_data_row}:F{last_data_row})",
				f"=SUM(G{first_data_row}:G{last_data_row})",
			]
		else:
			total_row_data += [0, 0]

		ws.append(total_row_data)
		total_row = ws.max_row
		for col in range(1, len(headers) + 1):
			cell      = ws.cell(row=total_row, column=col)
			cell.font = Font(bold=True)
			cell.fill = PatternFill("solid", start_color="D9E1F2")

	def apply_sheet_format(self, ws) -> None:
		"""Aplica formato final a la hoja: centra contenido y ajusta anchos de columna.

		Args:
			ws: Hoja de Excel (``Worksheet``) a formatear.
		"""
		for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
			for cell in row:
				cell.alignment = Alignment(horizontal="center")

		col_widths = [10, 16, 20, 14, 18, 24, 26]
		for i, width in enumerate(col_widths, 1):
			ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

	def save_trajectory_image(
		self,
		video_name: str,
		traj_x: list,
		traj_y: list,
		total_distance: float,
		first_frame: np.ndarray,
		output_dir: str,
	) -> None:
		"""Guarda una imagen PNG de la trayectoria completa sobre el primer frame.

		Args:
			video_name: Nombre del video sin extensión, usado en el nombre del archivo.
			traj_x: Lista de coordenadas x de la trayectoria.
			traj_y: Lista de coordenadas y de la trayectoria.
			total_distance: Distancia total recorrida en píxeles.
			first_frame: Primer frame del video como fondo (np.ndarray), o None.
			output_dir: Carpeta donde guardar el PNG.
		"""
		if first_frame is None or len(traj_x) <= 1:
			return

		background = first_frame.copy()
		trajectory = list(zip(traj_x, traj_y))

		for region in self.regions.regions:
			region.draw(background, (0, 255, 0), 2)

		for i in range(1, len(trajectory)):
			cv2.line(background, trajectory[i - 1], trajectory[i], (255, 0, 255), 2)

		cv2.putText(background, f"Distancia: {total_distance:.0f} px",
					(10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)

		img_path = os.path.join(output_dir, f"trajectory_{video_name}.png")
		cv2.imwrite(img_path, background)
		print(f"Imagen guardada en: {img_path}")

	def make_black_canvas(self, first_frame: np.ndarray) -> tuple:
		"""Crea un canvas negro del mismo tamaño que ``first_frame``.

		Args:
			first_frame: Frame de referencia para obtener las dimensiones.
				Si es None, se usa un tamaño genérico de 600×800.

		Returns:
			Tupla ``(canvas, height, width)`` donde ``canvas`` es un
			np.ndarray de ceros en BGR.
		"""
		if first_frame is not None:
			height, width = first_frame.shape[:2]
		else:
			height, width = 600, 800
		canvas = np.zeros((height, width, 3), dtype=np.uint8)
		return canvas, height, width

	def draw_arena_outline(self, canvas: np.ndarray) -> None:
		"""Dibuja el contorno del área experimental sobre el canvas.

		Implementación base vacía. Las subclases que necesiten dibujar un
		contorno específico (e.g. el círculo de la piscina en MorrisPool)
		deben sobreescribir este método.

		Args:
			canvas: Imagen BGR sobre la cual dibujar.
		"""
		pass

	def compute_density(
		self,
		all_trajectories: dict,
		height: int,
		width: int,
	) -> np.ndarray | None:
		"""Acumula la densidad de paso de todas las trayectorias y la normaliza.

		Dibuja líneas entre puntos consecutivos de cada trayectoria sobre una
		imagen de acumulación flotante, aplica un suavizado gaussiano y normaliza
		el resultado a ``[0, 255]``.

		Args:
			all_trajectories: ``{nombre_video: (traj_x, traj_y)}``.
			height: Alto de la imagen de densidad en píxeles.
			width: Ancho de la imagen de densidad en píxeles.

		Returns:
			Array uint8 normalizado ``[0, 255]``, o None si no hay trayectorias
			con más de un punto.
		"""
		density = np.zeros((height, width), dtype=np.float32)
		for traj_x, traj_y in all_trajectories.values():
			if len(traj_x) <= 1:
				continue
			trajectory = list(zip(traj_x, traj_y))
			for j in range(1, len(trajectory)):
				cv2.line(density, trajectory[j - 1], trajectory[j], color=1.0, thickness=1)

		if density.max() == 0:
			print("No hay trayectorias para generar heatmap.")
			return None

		density = cv2.GaussianBlur(density, (31, 31), 0)
		density_norm = cv2.normalize(density, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
		return density_norm

	def save_heatmap_image(
		self,
		all_trajectories: dict,
		first_frame: np.ndarray,
		output_dir: str,
	) -> None:
		"""Genera un mapa de calor combinando todas las trayectorias sobre fondo negro.

		Las zonas más transitadas aparecen en rojo/amarillo (caliente) y las
		menos transitadas en azul/verde (frío). El fondo negro indica sin actividad.

		Args:
			all_trajectories: ``{nombre_video: (traj_x, traj_y)}``.
			first_frame: Primer frame de cualquier video, usado para obtener
				las dimensiones del canvas. Puede ser None.
			output_dir: Carpeta donde guardar el PNG.
		"""
		canvas, height, width = self.make_black_canvas(first_frame)
		self.draw_arena_outline(canvas)

		density_norm = self.compute_density(all_trajectories, height, width)
		if density_norm is None:
			return

		heatmap_color = cv2.applyColorMap(density_norm, cv2.COLORMAP_JET)

		mask = density_norm > 0
		canvas[mask] = heatmap_color[mask]

		for region in self.regions.regions:
			region.draw(canvas, (255, 255, 255), 2)

		img_path = os.path.join(
			output_dir,
			f"heatmap_{self.mace_type}_{self.subject_id}_{self.treatment}.png",
		)
		cv2.imwrite(img_path, canvas)
		print(f"Mapa de calor guardado en: {img_path}")